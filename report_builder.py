"""
report_builder.py  —  Excel report generation for LCY3 AFM Dashboard
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1A237E")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
GRAND_FILL = PatternFill("solid", fgColor="E8EAF6")
GRAND_FONT = Font(color="1A237E", bold=True, size=10)
RED_FILL   = PatternFill("solid", fgColor="D22828")
ORG_FILL   = PatternFill("solid", fgColor="FF8C00")
GRN_FILL   = PatternFill("solid", fgColor="3CB44B")
WHT_FONT   = Font(color="FFFFFF", bold=True, size=10)
BLK_FONT   = Font(color="000000", bold=True, size=10)
THIN       = Side(style="thin", color="C5CAE9")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)

DEFAULT_THRESHOLD = 5
THRESHOLDS = {"Amnesty": 10, "Drive Lacking Capability": 10}
NO_THRESHOLD = ["Unreachable Charger"]


def _get_threshold(andon_type):
    if andon_type in NO_THRESHOLD:
        return None
    return THRESHOLDS.get(andon_type, DEFAULT_THRESHOLD)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _header_row(ws, values, row=1):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = BORDER


def _style_cell_border(ws, row, col):
    ws.cell(row=row, column=col).border = BORDER


def _colour_avg(cell, val, threshold):
    if val is None or (isinstance(val, float) and pd.isna(val)) or threshold is None:
        return
    if val > threshold * 1.5:
        cell.fill = RED_FILL; cell.font = WHT_FONT
    elif val > threshold:
        cell.fill = ORG_FILL; cell.font = BLK_FONT
    else:
        cell.fill = GRN_FILL; cell.font = WHT_FONT


# ── Sheet: Summary KPIs ───────────────────────────────────────────────────────
def _sheet_kpis(wb, fdf, uploaded_files, within_threshold_fn):
    ws = wb.create_sheet("Summary KPIs")
    within_pct = fdf.apply(within_threshold_fn, axis=1).mean() * 100
    eff = (fdf.groupby("Resolver").agg(n=("Resolve_Min", "count"), avg=("Resolve_Min", "mean"))
           .apply(lambda r: r["n"] / r["avg"] if r["avg"] > 0 else 0, axis=1).mean())

    _header_row(ws, ["Metric", "Value"], row=1)
    rows = [
        ("Total Andons",             len(fdf)),
        ("Avg Resolve Time (min)",   round(fdf["Resolve_Min"].mean(), 2)),
        ("Median Resolve Time (min)",round(fdf["Resolve_Min"].median(), 2)),
        ("% Within Threshold",       round(within_pct, 1)),
        ("Avg Efficiency Score",     round(eff, 1)),
        ("Files Uploaded",           len(uploaded_files)),
        ("Date Range",               f"{fdf['Date'].min()} → {fdf['Date'].max()}"),
    ]
    for r, (m, v) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=m).border = BORDER
        ws.cell(row=r, column=2, value=v).border = BORDER
    _auto_width(ws)


# ── Sheet: Leaderboard ────────────────────────────────────────────────────────
def _sheet_leaderboard(wb, fdf, within_threshold_fn):
    ws = wb.create_sheet("Leaderboard")
    lb = (fdf.groupby("Resolver")
          .agg(Total_Andons=("Resolve_Min", "count"), Avg_Time=("Resolve_Min", "mean"))
          .reset_index())
    lb["Avg_Time"]         = lb["Avg_Time"].round(2)
    lb["Efficiency"]       = (lb["Total_Andons"] / lb["Avg_Time"]).round(2)
    lb["Within_Threshold"] = fdf.groupby("Resolver").apply(
        lambda g: g.apply(within_threshold_fn, axis=1).mean() * 100
    ).round(1).values
    lb = lb.sort_values("Avg_Time").reset_index(drop=True)
    lb.index += 1

    headers = ["Rank", "Resolver", "Total Andons", "Avg Time (min)", "Efficiency Score", "% Within Threshold"]
    _header_row(ws, headers, row=1)

    for rank, row in lb.iterrows():
        r = rank + 1
        ws.cell(row=r, column=1, value=rank).border = BORDER
        ws.cell(row=r, column=2, value=row["Resolver"]).border = BORDER
        ws.cell(row=r, column=3, value=int(row["Total_Andons"])).border = BORDER
        avg_cell = ws.cell(row=r, column=4, value=row["Avg_Time"])
        avg_cell.border = BORDER
        _colour_avg(avg_cell, row["Avg_Time"], DEFAULT_THRESHOLD)
        ws.cell(row=r, column=5, value=row["Efficiency"]).border = BORDER
        ws.cell(row=r, column=6, value=round(row["Within_Threshold"], 1)).border = BORDER

    _auto_width(ws)


# ── Sheet: AFM Performance ────────────────────────────────────────────────────
def _sheet_afm_performance(wb, fdf):
    ws = wb.create_sheet("AFM Performance")
    andon_types   = sorted(fdf["Andon Type"].dropna().unique())
    all_resolvers = sorted(fdf["Resolver"].unique())

    # Build header
    headers = ["Resolver"]
    for at in andon_types:
        headers += [f"{at} — Count", f"{at} — Avg (min)"]
    headers += ["Total Count", "Total Avg (min)"]
    _header_row(ws, headers, row=1)

    cp = fdf.pivot_table(index="Resolver", columns="Andon Type",
                         values="Resolve_Min", aggfunc="count", fill_value=0).reindex(all_resolvers, fill_value=0)
    ap = fdf.pivot_table(index="Resolver", columns="Andon Type",
                         values="Resolve_Min", aggfunc="mean").round(2).reindex(all_resolvers)

    for r, res in enumerate(all_resolvers, 2):
        col = 1
        ws.cell(row=r, column=col, value=res).border = BORDER; col += 1
        for at in andon_types:
            cnt_val = int(cp.loc[res, at]) if at in cp.columns else 0
            avg_val = round(ap.loc[res, at], 2) if at in ap.columns and not pd.isna(ap.loc[res, at]) else None
            ws.cell(row=r, column=col, value=cnt_val).border = BORDER; col += 1
            avg_cell = ws.cell(row=r, column=col, value=avg_val)
            avg_cell.border = BORDER
            if avg_val:
                _colour_avg(avg_cell, avg_val, DEFAULT_THRESHOLD)
            col += 1
        total_cnt = int(cp.loc[res, andon_types].sum()) if len(andon_types) > 0 else 0
        total_avg = round(fdf[fdf["Resolver"] == res]["Resolve_Min"].mean(), 2)
        ws.cell(row=r, column=col, value=total_cnt).border = BORDER; col += 1
        avg_cell = ws.cell(row=r, column=col, value=total_avg)
        avg_cell.border = BORDER
        _colour_avg(avg_cell, total_avg, DEFAULT_THRESHOLD)

    # Grand total row
    gr = len(all_resolvers) + 2
    ws.cell(row=gr, column=1, value="Grand Total").fill = GRAND_FILL
    ws.cell(row=gr, column=1).font = GRAND_FONT
    ws.cell(row=gr, column=1).border = BORDER
    col = 2
    for at in andon_types:
        sub = fdf[fdf["Andon Type"] == at]
        ws.cell(row=gr, column=col, value=int(sub["Resolve_Min"].count())).fill = GRAND_FILL
        ws.cell(row=gr, column=col).font = GRAND_FONT
        ws.cell(row=gr, column=col).border = BORDER; col += 1
        ws.cell(row=gr, column=col, value=round(sub["Resolve_Min"].mean(), 2)).fill = GRAND_FILL
        ws.cell(row=gr, column=col).font = GRAND_FONT
        ws.cell(row=gr, column=col).border = BORDER; col += 1
    ws.cell(row=gr, column=col, value=int(fdf["Resolve_Min"].count())).fill = GRAND_FILL
    ws.cell(row=gr, column=col).font = GRAND_FONT
    ws.cell(row=gr, column=col).border = BORDER; col += 1
    ws.cell(row=gr, column=col, value=round(fdf["Resolve_Min"].mean(), 2)).fill = GRAND_FILL
    ws.cell(row=gr, column=col).font = GRAND_FONT
    ws.cell(row=gr, column=col).border = BORDER

    _auto_width(ws)


# ── Sheet: Andons by Type ─────────────────────────────────────────────────────
def _sheet_by_type(wb, fdf):
    ws = wb.create_sheet("Andons by Type")
    tc = fdf["Andon Type"].value_counts().reset_index()
    tc.columns = ["Andon Type", "Count"]
    tc["% of Total"]    = (tc["Count"] / tc["Count"].sum() * 100).round(1)
    tc["Avg Time (min)"] = tc["Andon Type"].map(fdf.groupby("Andon Type")["Resolve_Min"].mean().round(2))

    _header_row(ws, ["Andon Type", "Count", "% of Total", "Avg Time (min)"], row=1)
    for r, row in tc.iterrows():
        ws.cell(row=r+2, column=1, value=row["Andon Type"]).border = BORDER
        ws.cell(row=r+2, column=2, value=int(row["Count"])).border = BORDER
        ws.cell(row=r+2, column=3, value=row["% of Total"]).border = BORDER
        avg_cell = ws.cell(row=r+2, column=4, value=row["Avg Time (min)"])
        avg_cell.border = BORDER
        t = _get_threshold(row["Andon Type"])
        _colour_avg(avg_cell, row["Avg Time (min)"], t)
    _auto_width(ws)


# ── Sheet: Weekly Breakdown ───────────────────────────────────────────────────
def _sheet_weekly(wb, fdf):
    ws = wb.create_sheet("Weekly Breakdown")
    weeks_avail = sorted(fdf["Week"].dropna().unique(), reverse=True)

    headers = ["Andon Type"]
    for w in weeks_avail:
        headers += [f"Wk {w} — Andons", f"Wk {w} — Avg (min)"]
    headers += ["Total Andons", "Total Avg (min)"]
    _header_row(ws, headers, row=1)

    andon_types = sorted(fdf["Andon Type"].dropna().unique())
    for r, at in enumerate(andon_types, 2):
        col = 1
        ws.cell(row=r, column=col, value=at).border = BORDER; col += 1
        for w in weeks_avail:
            sub_w = fdf[(fdf["Week"] == w) & (fdf["Andon Type"] == at)]
            cnt = int(sub_w["Resolve_Min"].count())
            avg = round(sub_w["Resolve_Min"].mean(), 2) if cnt > 0 else None
            ws.cell(row=r, column=col, value=cnt).border = BORDER; col += 1
            avg_cell = ws.cell(row=r, column=col, value=avg)
            avg_cell.border = BORDER
            if avg:
                _colour_avg(avg_cell, avg, DEFAULT_THRESHOLD)
            col += 1
        sub_at = fdf[fdf["Andon Type"] == at]
        ws.cell(row=r, column=col, value=int(sub_at["Resolve_Min"].count())).border = BORDER; col += 1
        t_avg = round(sub_at["Resolve_Min"].mean(), 2)
        avg_cell = ws.cell(row=r, column=col, value=t_avg)
        avg_cell.border = BORDER
        _colour_avg(avg_cell, t_avg, _get_threshold(at))

    # Grand total
    gr = len(andon_types) + 2
    ws.cell(row=gr, column=1, value="Grand Total").fill = GRAND_FILL
    ws.cell(row=gr, column=1).font = GRAND_FONT
    ws.cell(row=gr, column=1).border = BORDER
    col = 2
    for w in weeks_avail:
        sub_w = fdf[fdf["Week"] == w]
        ws.cell(row=gr, column=col, value=int(sub_w["Resolve_Min"].count())).fill = GRAND_FILL
        ws.cell(row=gr, column=col).font = GRAND_FONT
        ws.cell(row=gr, column=col).border = BORDER; col += 1
        ws.cell(row=gr, column=col, value=round(sub_w["Resolve_Min"].mean(), 2)).fill = GRAND_FILL
        ws.cell(row=gr, column=col).font = GRAND_FONT
        ws.cell(row=gr, column=col).border = BORDER; col += 1
    ws.cell(row=gr, column=col, value=int(fdf["Resolve_Min"].count())).fill = GRAND_FILL
    ws.cell(row=gr, column=col).font = GRAND_FONT
    ws.cell(row=gr, column=col).border = BORDER; col += 1
    ws.cell(row=gr, column=col, value=round(fdf["Resolve_Min"].mean(), 2)).fill = GRAND_FILL
    ws.cell(row=gr, column=col).font = GRAND_FONT
    ws.cell(row=gr, column=col).border = BORDER

    _auto_width(ws)


# ── Sheet: Raw Data ───────────────────────────────────────────────────────────
def _sheet_raw(wb, fdf):
    ws = wb.create_sheet("Raw Data")
    cols = [c for c in fdf.columns if not c.startswith("_")]
    _header_row(ws, cols, row=1)
    for r, (_, row) in enumerate(fdf[cols].iterrows(), 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=str(val) if not isinstance(val, (int, float)) else val)
            cell.border = BORDER
    _auto_width(ws)


# ── Sheet: System vs Non-System ───────────────────────────────────────────────
def _sheet_sys_nonsys(wb, fdf):
    ws = wb.create_sheet("System vs Non-System")
    fdf2 = fdf.copy()
    fdf2["_rtype"] = fdf2["Resolver"].apply(lambda r: "System" if r == "System" else "Non-System")
    weeks_avail = sorted(fdf2["Week"].dropna().unique(), reverse=True)

    headers = ["Resolver Type"]
    for w in weeks_avail:
        headers += [f"Wk {w} — Andons", f"Wk {w} — Avg (min)"]
    headers += ["Total Andons", "Total Avg (min)", "% of Total"]
    _header_row(ws, headers, row=1)

    for r, rtype in enumerate(["System", "Non-System"], 2):
        sub = fdf2[fdf2["_rtype"] == rtype]
        col = 1
        ws.cell(row=r, column=col, value=rtype).border = BORDER; col += 1
        for w in weeks_avail:
            sub_w = sub[sub["Week"] == w]
            cnt = int(sub_w["Resolve_Min"].count())
            avg = round(sub_w["Resolve_Min"].mean(), 2) if cnt > 0 else None
            ws.cell(row=r, column=col, value=cnt).border = BORDER; col += 1
            ws.cell(row=r, column=col, value=avg).border = BORDER; col += 1
        total = int(sub["Resolve_Min"].count())
        total_avg = round(sub["Resolve_Min"].mean(), 2) if total > 0 else None
        pct = round(total / len(fdf2) * 100, 1)
        ws.cell(row=r, column=col, value=total).border = BORDER; col += 1
        ws.cell(row=r, column=col, value=total_avg).border = BORDER; col += 1
        ws.cell(row=r, column=col, value=pct).border = BORDER

    _auto_width(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def build_daily_report(fdf, uploaded_files, within_threshold_fn) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    _sheet_kpis(wb, fdf, uploaded_files, within_threshold_fn)
    _sheet_afm_performance(wb, fdf)
    _sheet_by_type(wb, fdf)
    _sheet_leaderboard(wb, fdf, within_threshold_fn)
    _sheet_raw(wb, fdf)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_weekly_report(fdf, uploaded_files, within_threshold_fn) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_kpis(wb, fdf, uploaded_files, within_threshold_fn)
    _sheet_weekly(wb, fdf)
    _sheet_afm_performance(wb, fdf)
    _sheet_sys_nonsys(wb, fdf)
    _sheet_leaderboard(wb, fdf, within_threshold_fn)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
